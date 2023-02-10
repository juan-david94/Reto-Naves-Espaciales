from openpyxl import load_workbook

# Se crea un clase padre en las que tiene como atributos comunes para todos los tipos de naves
# nombre y pais
class nave:
    def __init__(self, nombre, pais):
        self.nombre = nombre
        self.pais = pais

#Se crean clases heredadas con el fin de implementar las propias caractericas para cada tipo de nave
class vehiculoLanzadera(nave):
    def __init__(self, nombre, pais, peso, empuje, combustible, altura):
        super().__init__(nombre, pais)
        self.peso = peso
        self.empuje = empuje
        self.combustible = combustible
        self.altura = altura

# Se crea lista con los datos de las naves para almacenar en la base de datos (Archivo excel)
# En la hoja correspondiente segun el tipo de nave
        datos = [nombre, pais, peso, empuje, combustible, altura]
        archivo = "RegistroNaves3.xlsx"
        wb = load_workbook(archivo)
        hoja1 = wb['Vehiculos Lanzadera']
        wb.active = hoja1
        hoja1.append(datos)
        wb.save(archivo)

# Se definen metodos para visualizar caracteristicas de cada nave
    def visualizarLanzadera(self):
        print("Caracteristicas\n", "   ",
            "Nombre:", self.nombre,"\n    ", 
            "Pais", self.pais, "\n    ",
            "Peso:", self.peso,"toneladas \n    ",
            "Empuje:", self.empuje,"toneladas \n    ",
            "Combustible:",self.combustible,"\n    ",
            "Altura:",self.altura, "metros\n")


#Se crean clases heredadas con el fin de implementar las propias caractericas para cada tipo de nave
class navesNoTripuladas(nave):
    def __init__(self, nombre, pais, empuje, desplazamiento, combustible, funcionalidad):
        super().__init__(nombre, pais)
        self.empuje = empuje
        self.desplazamiento = desplazamiento
        self.combustible = combustible
        self.funcionalidad = funcionalidad

# Se crea lista con los datos de las naves para almacenar en la base de datos (Archivo excel)
# En la hoja correspondiente segun el tipo de nave
        datos = [nombre, pais, empuje, desplazamiento, combustible, funcionalidad]
        archivo = "RegistroNaves3.xlsx"
        wb = load_workbook(archivo)
        hoja1 = wb['Naves no tripuladas']
        wb.active = hoja1
        hoja1.append(datos)
        wb.save(archivo)

# Se definen metodos para visualizar caracteristicas de cada nave    
    def visualizarNotripuladas(self):
        if type(self.desplazamiento) == int:
            unidades = "Km\h"
        print("Caracteristicas\n", "   ",
            "Nombre:",self.nombre,"\n    ", 
            "Pais:", self.pais, "\n    ",
            "Empuje:", self.empuje,"Kilogramos \n    ",
            "Desplazamiento:", self.desplazamiento, unidades,"    ",
            "Combustible:", self.combustible,"\n    ",
            "Funcionalidad:", self.funcionalidad,"\n")

#Se crean clases heredadas con el fin de implementar las propias caractericas para cada tipo de nave        
class navesTripuladas(nave):
    def __init__(self, nombre, pais, peso, capacidad, distancia):
        super().__init__(nombre, pais)
        self.peso = peso
        self.capacidad = capacidad
        self.distancia = distancia

# Se crea lista con los datos de las naves para almacenar en la base de datos (Archivo excel)
# En la hoja correspondiente segun el tipo de nave
        datos = [nombre, pais, peso, capacidad, distancia]
        archivo = "RegistroNaves3.xlsx"
        wb = load_workbook(archivo)
        hoja1 = wb['Naves Tripuladas']
        wb.active = hoja1
        hoja1.append(datos)
        wb.save(archivo)

# Se definen metodos para visualizar caracteristicas de cada nave    
    def visualizarTripuladas(self):
        print("Caracteristicas\n", "   ",
            "Nombre:", self.nombre,"\n    ", 
            "Pais:", self.pais, "\n    ",
            "Peso:", self.peso, "toneladas\n    ",
            "Capacidad:", self.capacidad, "personas\n    ",
            "Distancia:", self.distancia,"km\n")


