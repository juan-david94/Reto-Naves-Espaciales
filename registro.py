import openpyxl

# Se crea funcion para crear el archivo donde se almacenan los datos de las naves registradas
def crear_archivo():
    wb = openpyxl.Workbook()

    hoja_lanzadera = wb.active
    hoja_lanzadera.title = "Vehiculos Lanzadera"
    hoja_noTripuladas = wb.create_sheet("Naves no tripuladas")
    hoja_tripuladas = wb.create_sheet("Naves Tripuladas")

    wb.save("RegistroNaves3.xlsx")
